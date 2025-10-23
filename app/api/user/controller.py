from flask import request, jsonify
from app.config.config import get_db_connection
from app.models.user import User
from werkzeug.utils import secure_filename
from io import BytesIO
from app import db
import openpyxl
import requests
import os
import csv


def create_users_table():
    """Create the users table if it doesn't exist."""
    conn = get_db_connection()
    cursor = conn.cursor()
    cursor.execute("""
        CREATE TABLE IF NOT EXISTS users (
            id INT AUTO_INCREMENT PRIMARY KEY,
            username VARCHAR(50) NOT NULL UNIQUE,
            email VARCHAR(120) NOT NULL UNIQUE,
            password VARCHAR(255) NOT NULL
        )
    """)
    conn.commit()
    cursor.close()
    conn.close()


def create_user():
    create_users_table()

    data = request.get_json()
    conn = get_db_connection()
    cursor = conn.cursor()
    sql = "INSERT INTO users (username, email, password) VALUES (%s, %s, %s)"
    values = (data['username'], data['email'], data['password'])
    cursor.execute(sql, values)
    conn.commit()
    cursor.close()
    conn.close()
    return jsonify({"message": "User created successfully"}), 201


# def create_user_bulk():

#     data = request.get_json()
#     file_url = data.get("file_url")

#     if not file_url:
#         return jsonify({"error": "Please provide 'file_url' in JSON body"}), 400

#     response = requests.get(file_url)
#     response.raise_for_status()

#     workbook = openpyxl.load_workbook(BytesIO(response.content))
#     sheet = workbook.active

#     for row in range(2, sheet.max_row + 1):
#         username = sheet.cell(row=row, column=1).value
#         email = sheet.cell(row=row, column=2).value
#         password = sheet.cell(row=row, column=3).value

#         if not (username and email and password):
#             continue 

#         existing_user = User.query.filter_by(email=email).first()


#         if existing_user:
#             existing_user = username
#             existing_user = password
#         else:
#             new_user = User(username=username, email=email, password=password)
#             db.session.add(new_user)

#     db.session.commit()
#     return jsonify({"message": "Bulk upload completed — existing users updated, new users added."}), 201

# UPLOAD_FOLDER = 'uploads'
# os.makedirs(UPLOAD_FOLDER, exist_ok=True)


def create_user_bulk():
    try:
        if 'file' not in request.files:
            return jsonify({"error": "No file part in the request"}), 400

        file = request.files['file']

        if file.filename == '':
            return jsonify({"error": "No file selected"}), 400

        filename = secure_filename(file.filename)
        ext = os.path.splitext(filename)[-1].lower()

        users_data = []

        if ext in [".xlsx", ".xls"]:
            workbook = openpyxl.load_workbook(BytesIO(file.read()))
            sheet = workbook.active
            for row in range(2, sheet.max_row + 1):
                username = sheet.cell(row=row, column=1).value
                email = sheet.cell(row=row, column=2).value
                password = sheet.cell(row=row, column=3).value
                if username and email and password:
                    users_data.append((username, email, password))

        elif ext == ".csv":
            decoded_content = file.read().decode('utf-8').splitlines()
            csv_reader = csv.reader(decoded_content)
            next(csv_reader, None)  # Skip header
            for row in csv_reader:
                if len(row) >= 3:
                    username, email, password = row[0], row[1], row[2]
                    if username and email and password:
                        users_data.append((username, email, password))
        else:
            return jsonify({"error": "Unsupported file type. Please upload .xlsx, .xls, or .csv"}), 400
        
        for username, email, password in users_data:
            existing_user = User.query.filter_by(email=email).first()
            if existing_user:
                existing_user.username = username
                existing_user.password = password
            else:
                new_user = User(username=username, email=email, password=password)
                db.session.add(new_user)

        db.session.commit()

        return jsonify({
            "message": f"Bulk upload completed — {len(users_data)} records processed successfully."
        }), 201

    except Exception as e:
        db.session.rollback()
        return jsonify({"error": str(e)}), 500

    # return jsonify({w
    #     "message": f"File '{filename}' uploaded successfully (not an Excel file).",
    #     "file_path": filepath
    # }), 200


# def create_user_bulk():
#     try:
#         file_path = 'C:/Users/saran/OneDrive/Desktop/e_commerce/excel_files/e_commerce_details.xlsx'

#         workbook = openpyxl.load_workbook(file_path)
#         sheet = workbook.active

#         for row in range(2, sheet.max_row + 1):
#             username = sheet.cell(row=row, column=1).value
#             email = sheet.cell(row=row, column=2).value
#             password = sheet.cell(row=row, column=3).value

#             if not (username and email and password):
#                 continue 
#             existing_user = User.query.filter_by(email=email).first()

#             if existing_user:
#                 existing_user = username
#                 existing_user = password
#             else:
#                 new_user = User(username=username, email=email, password=password)
#                 db.session.add(new_user)


#         db.session.commit()
#         return jsonify({"message": "Bulk upload completed — existing users updated, new users added."}), 201

#     except Exception :
#         db.session.rollback()
#         return jsonify({"error": str(Exception)}), 500

def get_users():

    conn = get_db_connection()
    cursor = conn.cursor(dictionary=True)
    cursor.execute("SELECT id, username, email FROM users")
    users = cursor.fetchall()
    cursor.close()
    conn.close()
    return jsonify(users)


def update_user(user_id):

    data = request.get_json()
    conn = get_db_connection()
    cursor = conn.cursor()
    sql = "UPDATE users SET username=%s, email=%s WHERE id=%s"
    values = (data.get("username"), data.get("email"), user_id)
    cursor.execute(sql, values)
    conn.commit()
    cursor.close()
    conn.close()
    return jsonify({"message": "User updated successfully"})


def delete_user(user_id):

    conn = get_db_connection()
    cursor = conn.cursor()
    sql = "DELETE FROM users WHERE id=%s"
    cursor.execute(sql, (user_id,))
    conn.commit()
    cursor.close()
    conn.close()
    return jsonify({"message": "User deleted successfully"})

